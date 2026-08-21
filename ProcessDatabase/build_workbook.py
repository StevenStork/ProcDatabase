#!/usr/bin/env python3
"""Build ProcessDatabase.xlsm sheet/table shell (VBA is imported from ../vba)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "ProcessDatabase.xlsm"

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill("solid", fgColor="D5E5F9")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
BOLD = Font(bold=True)


def style_header_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = THIN


def add_table(ws, name, ref):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def hide_grid(ws):
    ws.sheet_view.showGridLines = False


def build_home(ws):
    ws["A1"] = "Home"
    ws["A1"].font = BOLD
    headers = [
        (3, "Base Part Number"),
        (4, "Active Part"),
        (5, "Date"),
        (6, "Days"),
        (7, "Highlight"),
        (8, "Home FFA"),
        (9, "Factories"),
    ]
    for col, title in headers:
        ws.cell(5, col, title)
    style_header_row(ws, 5, 3, 9)
    ws.cell(6, 3, "")
    ws.cell(6, 4, False)
    for col in range(3, 10):
        ws.cell(6, col).border = THIN
        ws.cell(6, col).alignment = CENTER
    widths = {1: 8, 3: 22, 4: 12, 5: 12, 6: 10, 7: 12, 8: 24, 9: 24, 11: 18, 15: 18, 19: 18}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    hide_grid(ws)


def build_part_template(ws):
    ws["A1"] = "Part"
    ws["A1"].font = BOLD
    ws["C1"] = "Base Part Number"
    ws["C1"].font = BOLD
    ws["C2"] = ""
    ws["C8"] = "Home FFA"
    ws["C8"].font = BOLD
    ws["C8"].alignment = CENTER
    ws["C9"].fill = HEADER_FILL
    ws["C9"].alignment = CENTER
    ws["C9"].border = THIN
    ws["E8"] = "Dash"
    ws["F8"] = "Active"
    ws["G8"] = "Product Line"
    ws["H8"] = "Use"
    style_header_row(ws, 8, 5, 8)

    home_ffa_dv = DataValidation(
        type="list",
        formula1="OFFSET(References!$B$2,0,0,MAX(1,COUNTA(References!$B:$B)-1),1)",
        allow_blank=True,
        showDropDown=False,
    )
    home_ffa_dv.error = "Select an FFA from the References list."
    home_ffa_dv.errorTitle = "Home FFA"
    ws.add_data_validation(home_ffa_dv)
    home_ffa_dv.add(ws["C9"])

    ops_headers = [
        "Operation Sequence",
        "Operation Code",
        "Imported Process Hours",
        "Imported Average Executions",
        "Batch Size",
        "Export Process Hours",
        "Export Average Executions",
        "Equipment Type",
        "Use Export Hours",
        "Use Export Executions",
        "Process Hours",
        "Average Executions",
        "Average HPUs",
        "FFA",
    ]
    for offset, title in enumerate(ops_headers):
        cell = ws.cell(8, 13 + offset, title)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = THIN
        cell.fill = HEADER_FILL

    # One starter data row for the ops range.
    for col in range(13, 27):
        cell = ws.cell(9, col, None)
        cell.border = THIN
        cell.alignment = CENTER
    ws["W9"] = '=IF(AND(R9<>"",U9=TRUE),R9,O9)'
    ws["X9"] = '=IF(AND(S9<>"",V9=TRUE),S9,P9)'
    ws["Y9"] = '=IF(OR(Q9="",W9="",X9=""),"",(W9*X9)/Q9)'
    ws["U9"] = False
    ws["V9"] = False
    for col in ("O", "P", "Q", "R", "S", "W", "X", "Y"):
        ws[f"{col}9"].number_format = "0.00"

    for letter, width in (
        ("C", 16),
        ("D", 8),
        ("E", 12),
        ("F", 8),
        ("G", 18),
        ("H", 8),
        ("M", 12),
        ("N", 10),
        ("T", 16),
        ("Z", 12),
    ):
        ws.column_dimensions[letter].width = width
    hide_grid(ws)


def build_assembly_standards(ws):
    ws["A1"] = "Standards"
    ws["A1"].font = BOLD
    headers = ["ASSEMBLY NO", "OPER SEQ", "FFA", "RUN TIME (HOURS)"]
    for col, title in enumerate(headers, start=1):
        ws.cell(5, col, title)
    style_header_row(ws, 5, 1, 4)
    for col in range(1, 5):
        ws.cell(6, col, None).border = THIN
    add_table(ws, "AssyStndTbl", "A5:D6")
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22


def build_references(ws):
    ws["A1"] = "Refs"
    ws["A1"].font = BOLD
    headers = {2: "FFA", 3: "Factory", 4: "Product Line", 5: "Equipment", 6: "Owning FFAs"}
    for col, title in headers.items():
        ws.cell(1, col, title)
        ws.cell(1, col).font = BOLD
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.sheet_state = "veryHidden"


def build_data(ws):
    ws["A1"] = "Data"
    ws["A1"].font = BOLD
    ws["A2"] = "HomeListHash"
    ws["B2"] = ""
    ws["A3"] = "ExportOpsHash"
    ws["B3"] = ""
    ws["A4"] = "RefsDirty"
    ws["B4"] = True
    ws["A5"] = "UiSchema"
    ws["B5"] = "1"

    part_headers = [
        "Base Part",
        "Active",
        "Home FFA",
        "Factories",
        "Product Lines",
        "Dashes",
        "UiSchema",
        "ListSig",
        "OpsDirty",
        "OpsRowCount",
        "SheetName",
    ]
    for col, title in enumerate(part_headers, start=1):
        ws.cell(8, col, title)
    style_header_row(ws, 8, 1, len(part_headers))
    for col in range(1, len(part_headers) + 1):
        ws.cell(9, col, None).border = THIN
    add_table(ws, "tblParts", f"A8:{get_column_letter(len(part_headers))}9")

    ops_headers = [
        "Part Number",
        "Operation Sequence",
        "Operation Code",
        "Imported Process Hours",
        "Imported Average Executions",
        "Batch Size",
        "Export Process Hours",
        "Export Average Executions",
        "Equipment Type",
        "Use Export Hours",
        "Use Export Executions",
        "Process Hours",
        "Average Executions",
        "Average HPUs",
        "FFA",
    ]
    ops_start_col = 13
    ops_end_col = ops_start_col + len(ops_headers) - 1
    for col, title in enumerate(ops_headers):
        ws.cell(8, ops_start_col + col, title)
    style_header_row(ws, 8, ops_start_col, ops_end_col)
    for col in range(ops_start_col, ops_end_col + 1):
        ws.cell(9, col, None).border = THIN
    add_table(
        ws,
        "tblOperations",
        f"{get_column_letter(ops_start_col)}8:{get_column_letter(ops_end_col)}9",
    )
    for col in range(1, ops_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.sheet_state = "veryHidden"


def main():
    wb = Workbook()
    home = wb.active
    home.title = "Home"
    build_home(home)

    part = wb.create_sheet("Part Number Template")
    build_part_template(part)

    standards = wb.create_sheet("Assembly Standards")
    build_assembly_standards(standards)

    refs = wb.create_sheet("References")
    build_references(refs)

    data = wb.create_sheet("Data")
    build_data(data)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
